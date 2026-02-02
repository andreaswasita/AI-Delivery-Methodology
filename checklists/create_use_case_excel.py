#!/usr/bin/env python3
"""
Create Excel spreadsheet for Use Case Business Change Checklist
Converts the comprehensive use case checklist into an interactive Excel workbook
"""

import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

def parse_use_case_checklist(filepath):
    """Parse the use case business change checklist markdown file"""
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    
    items = []
    current_section = ""
    current_subsection = ""
    
    lines = content.split('\n')
    
    for line in lines:
        # Main section headers (## SECTION X:)
        if line.startswith('## SECTION '):
            current_section = line.replace('## ', '').strip()
            current_subsection = ""
            continue
        
        # Subsection headers (### )
        if line.startswith('### '):
            current_subsection = line.replace('### ', '').strip()
            continue
        
        # Checklist items (- [ ])
        if re.match(r'^\s*-\s*\[\s*\]', line):
            # Extract the item text
            item_text = re.sub(r'^\s*-\s*\[\s*\]\s*', '', line)
            # Remove markdown formatting
            item_text = re.sub(r'\*\*(.*?)\*\*', r'\1', item_text)  # Bold
            item_text = re.sub(r'\*(.*?)\*', r'\1', item_text)  # Italic
            item_text = item_text.strip()
            
            # Extract IDs if present (e.g., BC-01, EC-01, TE-01)
            id_match = re.match(r'^([A-Z]+-\d+):', item_text)
            item_id = id_match.group(1) if id_match else ""
            
            if item_text:
                items.append({
                    'section': current_section,
                    'subsection': current_subsection,
                    'item_id': item_id,
                    'item': item_text,
                    'status': '',
                    'owner': '',
                    'due_date': '',
                    'notes': ''
                })
    
    return items

def create_summary_sheet(wb):
    """Create summary/instructions sheet"""
    ws = wb.create_sheet("Instructions", 0)
    
    # Define styles
    title_font = Font(name='Calibri', size=16, bold=True, color='366092')
    header_font = Font(name='Calibri', size=12, bold=True)
    normal_font = Font(name='Calibri', size=10)
    
    # Title
    ws['A1'] = 'Use Case Business Change & Value Realization Checklist'
    ws['A1'].font = title_font
    ws.merge_cells('A1:E1')
    
    # Instructions
    ws['A3'] = 'Purpose:'
    ws['A3'].font = header_font
    ws['A4'] = 'This checklist ensures systematic identification of business process changes, benefits dependencies, and value realization for each AI use case.'
    ws['A4'].font = normal_font
    ws['A4'].alignment = Alignment(wrap_text=True)
    ws.merge_cells('A4:E4')
    
    ws['A6'] = 'How to Use:'
    ws['A6'].font = header_font
    
    instructions = [
        '1. Complete one checklist per AI use case',
        '2. Work through all 11 sections systematically',
        '3. Update Status column as you progress (Not Started → In Progress → Completed)',
        '4. Assign owners for each item',
        '5. Set due dates for accountability',
        '6. Use Notes column for additional details',
        '7. Review completion before proceeding to implementation',
        '',
        'Status Values:',
        '  • Not Started - Item not yet begun',
        '  • In Progress - Currently working on this item',
        '  • Completed - Item finished and validated',
        '  • Blocked - Item cannot proceed (document blocker in Notes)',
        '  • N/A - Not applicable to this use case',
        '',
        'Critical Success Factor:',
        'Technology enables, but business change delivers value. Ensure all business changes are identified and planned.',
        '',
        'Related Resources:',
        '  • Benefits Dependency Network Template',
        '  • Business Process Change Analysis Framework',
        '  • Quick Reference Guide',
        '  • ROI Calculator',
    ]
    
    row = 7
    for instruction in instructions:
        ws[f'A{row}'] = instruction
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws.merge_cells(f'A{row}:E{row}')
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 80
    
    return ws

def create_checklist_sheet(wb, items, sheet_name):
    """Create main checklist sheet"""
    ws = wb.create_sheet(sheet_name)
    
    # Define styles
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    section_font = Font(name='Calibri', size=10, bold=True)
    section_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    cell_font = Font(name='Calibri', size=10)
    cell_alignment = Alignment(vertical='top', wrap_text=True)
    
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Define column headers
    headers = ['Section', 'Subsection', 'Item ID', 'Task/Item', 'Status', 'Owner', 'Due Date', 'Notes']
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Set column widths
    column_widths = [30, 25, 10, 60, 15, 20, 12, 35]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Write data
    current_section = ""
    row_num = 2
    
    for item in items:
        # Add section separator row if section changed
        if item['section'] != current_section:
            current_section = item['section']
            # Section header row
            ws.merge_cells(f'A{row_num}:H{row_num}')
            cell = ws.cell(row=row_num, column=1)
            cell.value = current_section
            cell.font = section_font
            cell.fill = section_fill
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = border
            row_num += 1
        
        # Write item data
        ws.cell(row=row_num, column=1, value=item['section']).font = cell_font
        ws.cell(row=row_num, column=2, value=item['subsection']).font = cell_font
        ws.cell(row=row_num, column=3, value=item['item_id']).font = cell_font
        ws.cell(row=row_num, column=4, value=item['item']).font = cell_font
        ws.cell(row=row_num, column=5, value=item['status']).font = cell_font
        ws.cell(row=row_num, column=6, value=item['owner']).font = cell_font
        ws.cell(row=row_num, column=7, value=item['due_date']).font = cell_font
        ws.cell(row=row_num, column=8, value=item['notes']).font = cell_font
        
        # Apply alignment and borders
        for col_num in range(1, 9):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = cell_alignment
            cell.border = border
        
        row_num += 1
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Add data validation for Status column
    status_validation = DataValidation(
        type="list",
        formula1='"Not Started,In Progress,Completed,Blocked,N/A"',
        allow_blank=True
    )
    status_validation.add(f'E2:E{row_num}')
    ws.add_data_validation(status_validation)
    
    return ws

def create_npv_calculator_sheet(wb):
    """Create NPV calculator sheet"""
    ws = wb.create_sheet("NPV Calculator")
    
    # Define styles
    header_font = Font(name='Calibri', size=12, bold=True, color='366092')
    label_font = Font(name='Calibri', size=10, bold=True)
    normal_font = Font(name='Calibri', size=10)
    currency_format = '#,##0'
    percent_format = '0.0%'
    
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Title
    ws['A1'] = 'NPV Calculator - 5 Year Projection'
    ws['A1'].font = header_font
    ws.merge_cells('A1:H1')
    
    # Use Case Info
    ws['A3'] = 'Use Case ID:'
    ws['A3'].font = label_font
    ws['B3'] = '[Enter UC-XXX]'
    
    ws['A4'] = 'Use Case Name:'
    ws['A4'].font = label_font
    ws['B4'] = '[Enter name]'
    ws.merge_cells('B4:E4')
    
    ws['A5'] = 'Discount Rate (%):'
    ws['A5'].font = label_font
    ws['B5'] = 0.10
    ws['B5'].number_format = percent_format
    
    # Implementation Costs (Year 0)
    ws['A7'] = 'IMPLEMENTATION COSTS (Year 0)'
    ws['A7'].font = header_font
    
    cost_items = [
        'Technology licenses & subscriptions',
        'Azure infrastructure',
        'Data preparation & labeling',
        'AI model development',
        'Integration & APIs',
        'Training & change management',
        'External consultants',
        'Internal project team',
        'TOTAL IMPLEMENTATION COST'
    ]
    
    row = 8
    for item in cost_items:
        ws[f'A{row}'] = item
        ws[f'A{row}'].font = label_font if 'TOTAL' in item else normal_font
        ws[f'B{row}'] = 0 if 'TOTAL' not in item else f'=SUM(B8:B{row-1})'
        ws[f'B{row}'].number_format = currency_format
        ws[f'B{row}'].border = border
        row += 1
    
    # Annual Operating Costs
    ws[f'A{row+1}'] = 'ANNUAL OPERATING COSTS (Years 1-5)'
    ws[f'A{row+1}'].font = header_font
    
    operating_costs = [
        'Azure consumption',
        'Licenses & subscriptions',
        'Maintenance & support',
        'Model retraining & updates',
        'Operations team',
        'TOTAL ANNUAL OPERATING COST'
    ]
    
    row += 2
    start_op_row = row
    for item in operating_costs:
        ws[f'A{row}'] = item
        ws[f'A{row}'].font = label_font if 'TOTAL' in item else normal_font
        ws[f'B{row}'] = 0 if 'TOTAL' not in item else f'=SUM(B{start_op_row}:B{row-1})'
        ws[f'B{row}'].number_format = currency_format
        ws[f'B{row}'].border = border
        row += 1
    
    # Annual Benefits
    ws[f'A{row+1}'] = 'ANNUAL BENEFITS (Years 1-5)'
    ws[f'A{row+1}'].font = header_font
    
    benefits = [
        'Cost savings',
        'Revenue increase',
        'Risk mitigation value',
        'Productivity gains (monetized)',
        'TOTAL ANNUAL BENEFITS'
    ]
    
    row += 2
    start_benefit_row = row
    for item in benefits:
        ws[f'A{row}'] = item
        ws[f'A{row}'].font = label_font if 'TOTAL' in item else normal_font
        ws[f'B{row}'] = 0 if 'TOTAL' not in item else f'=SUM(B{start_benefit_row}:B{row-1})'
        ws[f'B{row}'].number_format = currency_format
        ws[f'B{row}'].border = border
        row += 1
    
    # NPV Calculation Table
    ws[f'A{row+2}'] = 'NPV CALCULATION'
    ws[f'A{row+2}'].font = header_font
    
    row += 3
    npv_headers = ['Year', 'Investment', 'Benefits', 'Net Cash Flow', 'Discount Factor', 'PV of Cash Flow']
    for col_num, header in enumerate(npv_headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.font = label_font
        cell.border = border
        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    # NPV calculation rows (Years 0-5)
    impl_cost_cell = 'B16'  # Reference to total implementation cost
    annual_op_cell = 'B23'  # Reference to annual operating cost
    annual_benefit_cell = 'B29'  # Reference to annual benefits
    discount_rate_cell = 'B5'
    
    row += 1
    for year in range(6):
        ws[f'A{row}'] = year
        
        if year == 0:
            ws[f'B{row}'] = f'={impl_cost_cell}'
            ws[f'C{row}'] = 0
        else:
            ws[f'B{row}'] = f'={annual_op_cell}'
            ws[f'C{row}'] = f'={annual_benefit_cell}'
        
        ws[f'D{row}'] = f'=C{row}-B{row}'
        ws[f'E{row}'] = f'=1/((1+{discount_rate_cell})^A{row})'
        ws[f'F{row}'] = f'=D{row}*E{row}'
        
        for col in ['B', 'C', 'D', 'F']:
            ws[f'{col}{row}'].number_format = currency_format
            ws[f'{col}{row}'].border = border
        ws[f'E{row}'].number_format = '0.000'
        ws[f'E{row}'].border = border
        
        row += 1
    
    # NPV Total
    ws[f'A{row}'] = 'NPV'
    ws[f'A{row}'].font = label_font
    ws[f'F{row}'] = f'=SUM(F{row-6}:F{row-1})'
    ws[f'F{row}'].number_format = currency_format
    ws[f'F{row}'].font = label_font
    ws[f'F{row}'].fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    ws[f'F{row}'].border = border
    
    # Financial metrics
    row += 2
    ws[f'A{row}'] = 'FINANCIAL METRICS'
    ws[f'A{row}'].font = header_font
    
    row += 1
    ws[f'A{row}'] = 'NPV:'
    ws[f'B{row}'] = f'=F{row-2}'
    ws[f'B{row}'].number_format = currency_format
    
    row += 1
    ws[f'A{row}'] = 'Decision:'
    ws[f'B{row}'] = f'=IF(B{row-1}>0,"✅ Proceed - Positive NPV","❌ Reject - Negative NPV")'
    
    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 20
    
    return ws

def create_bdn_tracker_sheet(wb):
    """Create Benefits Dependency Network tracker sheet"""
    ws = wb.create_sheet("Benefits Dependency Network")
    
    # Define styles
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    cell_font = Font(name='Calibri', size=10)
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Create sections for each BDN layer
    sections = [
        ('Business Drivers', ['ID', 'Business Driver', 'Description', 'Impact Level']),
        ('Investment Objectives', ['ID', 'Investment Objective', 'Success Criteria', 'Business Driver Link']),
        ('Benefits', ['ID', 'Benefit', 'Measurement', 'Baseline', 'Target', 'Investment Objective Link', 'Owner']),
        ('Business Changes', ['ID', 'Business Change', 'Description', 'Impacted Stakeholders', 'Benefit Link', 'Change Readiness']),
        ('Enabling Changes', ['ID', 'Enabling Change', 'Description', 'Business Change Link', 'Owner', 'Status']),
        ('IT Enablers', ['ID', 'Technology', 'Purpose', 'Enabling Change Link', 'Azure Service', 'Cost Estimate'])
    ]
    
    row = 1
    for section_title, headers in sections:
        # Section title
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws.cell(row=row, column=1)
        cell.value = section_title
        cell.font = Font(name='Calibri', size=12, bold=True, color='366092')
        row += 1
        
        # Headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        row += 1
        
        # Add 5 empty rows for data entry
        for _ in range(5):
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col_num)
                cell.font = cell_font
                cell.border = border
            row += 1
        
        row += 1  # Gap between sections
    
    # Set column widths
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 25
    
    return ws

def main():
    """Create the Use Case Business Change Checklist Excel file"""
    # Get current directory
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Input and output paths
    md_filepath = os.path.join(script_dir, 'use-case-business-change-checklist.md')
    excel_filepath = os.path.join(script_dir, 'Use-Case-Business-Change-Checklist.xlsx')
    
    print("Creating Use Case Business Change Checklist Excel file...")
    print(f"Reading from: {md_filepath}")
    
    # Check if markdown file exists
    if not os.path.exists(md_filepath):
        print(f"ERROR: File not found: {md_filepath}")
        return
    
    # Parse the checklist
    items = parse_use_case_checklist(md_filepath)
    print(f"Parsed {len(items)} checklist items")
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create sheets
    print("Creating Instructions sheet...")
    create_summary_sheet(wb)
    
    print("Creating Checklist sheet...")
    create_checklist_sheet(wb, items, "Checklist")
    
    print("Creating NPV Calculator sheet...")
    create_npv_calculator_sheet(wb)
    
    print("Creating Benefits Dependency Network sheet...")
    create_bdn_tracker_sheet(wb)
    
    # Save workbook
    wb.save(excel_filepath)
    print(f"\n✅ SUCCESS! Created: {excel_filepath}")
    print(f"   - Instructions sheet")
    print(f"   - Checklist sheet ({len(items)} items)")
    print(f"   - NPV Calculator sheet")
    print(f"   - Benefits Dependency Network sheet")

if __name__ == '__main__':
    main()
