#!/usr/bin/env python3
"""
Create individual Excel files for each phase checklist
Each checklist markdown file becomes a separate Excel workbook
"""

import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def parse_checklist(filepath):
    """Parse markdown checklist and extract items"""
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    
    items = []
    current_section = ""
    current_subsection = ""
    
    lines = content.split('\n')
    
    for line in lines:
        # Section headers (## )
        if line.startswith('## ') and not line.startswith('###'):
            current_section = line.replace('## ', '').strip()
            current_subsection = ""
            continue
        
        # Subsection headers (### )
        if line.startswith('### '):
            current_subsection = line.replace('### ', '').strip()
            continue
        
        # Checklist items (- [ ])
        if re.match(r'^\s*-\s*\[\s*\]', line):
            # Extract the item text
            item_text = re.sub(r'^\s*-\s*\[\s*\]\s*', '', line)
            # Remove markdown formatting
            item_text = re.sub(r'\*\*(.*?)\*\*', r'\1', item_text)  # Bold
            item_text = re.sub(r'\*(.*?)\*', r'\1', item_text)  # Italic
            item_text = item_text.strip()
            
            if item_text:
                items.append({
                    'section': current_section,
                    'subsection': current_subsection,
                    'item': item_text,
                    'status': '',
                    'notes': '',
                    'owner': '',
                    'due_date': ''
                })
    
    return items

def create_excel_for_checklist(md_filepath, excel_filepath):
    """Create Excel workbook for a single checklist"""
    print(f"Processing: {md_filepath}")
    
    # Parse the markdown checklist
    items = parse_checklist(md_filepath)
    
    if not items:
        print(f"  Warning: No checklist items found in {md_filepath}")
        return
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    
    # Get checklist name from filename
    checklist_name = os.path.basename(md_filepath).replace('-checklist.md', '').replace('-', ' ').title()
    ws.title = checklist_name[:31]  # Excel sheet name limit
    
    # Define styles
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_font = Font(name='Calibri', size=10)
    cell_alignment = Alignment(vertical='top', wrap_text=True)
    
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Define column headers
    headers = ['Phase/Section', 'Subsection', 'Task/Item', 'Status', 'Owner', 'Due Date', 'Notes']
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Set column widths
    column_widths = [25, 25, 50, 12, 20, 12, 30]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Write data
    for row_num, item in enumerate(items, 2):
        ws.cell(row=row_num, column=1, value=item['section']).font = cell_font
        ws.cell(row=row_num, column=2, value=item['subsection']).font = cell_font
        ws.cell(row=row_num, column=3, value=item['item']).font = cell_font
        ws.cell(row=row_num, column=4, value=item['status']).font = cell_font
        ws.cell(row=row_num, column=5, value=item['owner']).font = cell_font
        ws.cell(row=row_num, column=6, value=item['due_date']).font = cell_font
        ws.cell(row=row_num, column=7, value=item['notes']).font = cell_font
        
        # Apply alignment and borders
        for col_num in range(1, 8):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = cell_alignment
            cell.border = border
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Add data validation for Status column
    from openpyxl.worksheet.datavalidation import DataValidation
    status_validation = DataValidation(
        type="list",
        formula1='"Not Started,In Progress,Completed,Blocked,N/A"',
        allow_blank=True
    )
    status_validation.add(f'D2:D{len(items) + 1}')
    ws.add_data_validation(status_validation)
    
    # Save workbook
    wb.save(excel_filepath)
    print(f"  Created: {excel_filepath} ({len(items)} items)")

def main():
    """Process all phase checklist files"""
    # Get current directory
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Define phase checklist files (in order)
    phase_checklists = [
        '00-presales-discovery-checklist.md',
        '01-mobilisation-checklist.md',
        '02-hackathons-checklist.md',
        '03-setup-platform-checklist.md',
        '04-build-phase-checklist.md',
        '05-integrate-phase-checklist.md',
        '06-test-evaluate-phase-checklist.md',
        '07-prepare-deploy-phase-checklist.md',
        '08-operate-care-phase-checklist.md'
    ]
    
    print("Creating individual Excel files for phase checklists...")
    print("=" * 70)
    
    created_count = 0
    
    for md_file in phase_checklists:
        md_filepath = os.path.join(script_dir, md_file)
        
        if os.path.exists(md_filepath):
            # Create Excel filename
            excel_file = md_file.replace('.md', '.xlsx')
            excel_filepath = os.path.join(script_dir, excel_file)
            
            try:
                create_excel_for_checklist(md_filepath, excel_filepath)
                created_count += 1
            except Exception as e:
                print(f"  Error processing {md_file}: {str(e)}")
        else:
            print(f"  Warning: File not found - {md_file}")
    
    print("=" * 70)
    print(f"Complete! Created {created_count} Excel files")
    
    # Also check for business envisioning workshop checklist
    business_checklist = 'business-envisioning-workshop-checklist.md'
    business_filepath = os.path.join(script_dir, business_checklist)
    
    if os.path.exists(business_filepath):
        print("\nBonus: Creating Excel for business envisioning workshop...")
        excel_filepath = os.path.join(script_dir, business_checklist.replace('.md', '.xlsx'))
        try:
            create_excel_for_checklist(business_filepath, excel_filepath)
            print("Business envisioning workshop Excel created!")
        except Exception as e:
            print(f"Error: {str(e)}")

if __name__ == '__main__':
    main()
