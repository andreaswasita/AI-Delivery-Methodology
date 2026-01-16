#!/usr/bin/env python3
"""
Convert all markdown checklists to Excel workbook
Each checklist becomes a separate sheet with structured columns
"""

import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def parse_checklist(filepath):
    """Parse markdown checklist and extract items"""
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    
    items = []
    current_section = ""
    current_subsection = ""
    
    lines = content.split('\n')
    
    for line in lines:
        # Section headers (## )
        if line.startswith('## ') and not line.startswith('###'):
            current_section = line.replace('## ', '').strip()
            current_subsection = ""
            continue
        
        # Subsection headers (### )
        if line.startswith('### '):
            current_subsection = line.replace('### ', '').strip()
            continue
        
        # Checklist items (- [ ])
        if re.match(r'^\s*-\s*\[\s*\]', line):
            # Extract the item text
            item_text = re.sub(r'^\s*-\s*\[\s*\]\s*', '', line)
            # Remove markdown formatting
            item_text = re.sub(r'\*\*(.*?)\*\*', r'\1', item_text)  # Bold
            item_text = re.sub(r'\*(.*?)\*', r'\1', item_text)  # Italic
            item_text = item_text.strip()
            
            if item_text:
                items.append({
                    'section': current_section,
                    'subsection': current_subsection,
                    'item': item_text,
                    'status': '',
                    'date_completed': '',
                    'owner': '',
                    'notes': ''
                })
    
    return items

def create_excel_sheet(ws, checklist_name, items):
    """Create formatted Excel sheet for a checklist"""
    
    # Set sheet title
    ws.title = checklist_name[:31]  # Excel sheet name max 31 chars
    
    # Define styles
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    section_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Define columns
    columns = [
        ('Section', 25),
        ('Subsection', 25),
        ('Checklist Item', 60),
        ('Status', 12),
        ('Date Completed', 15),
        ('Owner', 20),
        ('Notes', 40)
    ]
    
    # Set column widths and headers
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = col_width
        
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Freeze top row
    ws.freeze_panes = 'A2'
    
    # Add data
    current_row = 2
    previous_section = ""
    
    for item in items:
        # Add section row if section changed
        if item['section'] != previous_section and item['section']:
            cell = ws.cell(row=current_row, column=1)
            cell.value = item['section']
            cell.fill = section_fill
            cell.font = section_font
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Merge across columns for section header
            ws.merge_cells(f'A{current_row}:G{current_row}')
            
            # Apply border to merged cell
            for col in range(1, 8):
                ws.cell(row=current_row, column=col).border = border
            
            current_row += 1
            previous_section = item['section']
        
        # Add checklist item
        ws.cell(row=current_row, column=1).value = item['section']
        ws.cell(row=current_row, column=2).value = item['subsection']
        ws.cell(row=current_row, column=3).value = item['item']
        ws.cell(row=current_row, column=4).value = ''  # Status
        ws.cell(row=current_row, column=5).value = ''  # Date Completed
        ws.cell(row=current_row, column=6).value = ''  # Owner
        ws.cell(row=current_row, column=7).value = ''  # Notes
        
        # Apply formatting
        for col in range(1, 8):
            cell = ws.cell(row=current_row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        current_row += 1
    
    # Add data validation for Status column
    from openpyxl.worksheet.datavalidation import DataValidation
    
    dv = DataValidation(
        type="list",
        formula1='"Not Started,In Progress,Completed,N/A"',
        allow_blank=True
    )
    dv.error = 'Please select from the list'
    dv.errorTitle = 'Invalid Status'
    
    ws.add_data_validation(dv)
    dv.add(f'D2:D{current_row}')

def main():
    # Get all checklist files
    checklist_dir = os.path.dirname(os.path.abspath(__file__))
    checklist_files = [
        f for f in os.listdir(checklist_dir)
        if f.endswith('-checklist.md')
    ]
    
    if not checklist_files:
        print("No checklist files found!")
        return
    
    # Sort files for consistent ordering
    checklist_files.sort()
    
    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Process each checklist
    for checklist_file in checklist_files:
        print(f"Processing {checklist_file}...")
        
        filepath = os.path.join(checklist_dir, checklist_file)
        checklist_name = checklist_file.replace('-checklist.md', '').replace('-', ' ').title()
        
        # Parse checklist
        items = parse_checklist(filepath)
        
        if items:
            # Create sheet
            ws = wb.create_sheet(title=checklist_name[:31])
            create_excel_sheet(ws, checklist_name, items)
            print(f"  ✓ Created sheet '{checklist_name}' with {len(items)} items")
        else:
            print(f"  ⚠ No checklist items found in {checklist_file}")
    
    # Create summary sheet
    ws_summary = wb.create_sheet(title="Overview", index=0)
    
    # Summary header
    ws_summary['A1'] = 'AI Delivery Methodology - Checklists'
    ws_summary['A1'].font = Font(bold=True, size=14, color="0066CC")
    ws_summary['A1'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws_summary['A3'] = 'Checklist Name'
    ws_summary['B3'] = 'Total Items'
    ws_summary['C3'] = 'Description'
    
    for col in ['A', 'B', 'C']:
        ws_summary[f'{col}3'].font = Font(bold=True)
        ws_summary[f'{col}3'].fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        ws_summary[f'{col}3'].font = Font(bold=True, color="FFFFFF")
    
    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.column_dimensions['C'].width = 60
    
    # Add summary data
    descriptions = {
        'Build Phase': 'Build production-quality software with comprehensive testing and deployment',
        'Business Envisioning': 'Workshop preparation and execution for AI use case identification',
        'Discovery': 'Understand business problem, assess data, and validate AI feasibility',
        'Hackathons Prototype': 'Rapid prototyping and hackathon execution checklist',
        'Integrate Phase': 'Post-deployment integration with enterprise systems and validation',
        'Mobilisation': 'Project setup, team mobilization, and governance establishment',
        'Operate Care Phase': 'Ongoing operations, monitoring, and continuous improvement',
        'Prepare Deploy Phase': 'Production deployment preparation and execution',
        'Setup Platform': 'Azure environment and MLOps platform setup',
        'Test Evaluate Phase': 'Comprehensive testing and model evaluation'
    }
    
    row = 4
    for sheet in wb.sheetnames:
        if sheet != 'Overview':
            ws_summary[f'A{row}'] = sheet
            ws_summary[f'A{row}'].hyperlink = f"#{sheet}!A1"
            ws_summary[f'A{row}'].style = "Hyperlink"
            
            # Count items in sheet
            target_sheet = wb[sheet]
            item_count = target_sheet.max_row - 1  # Exclude header
            ws_summary[f'B{row}'] = item_count
            ws_summary[f'B{row}'].alignment = Alignment(horizontal='center')
            
            ws_summary[f'C{row}'] = descriptions.get(sheet, '')
            row += 1
    
    # Save workbook
    output_file = os.path.join(checklist_dir, 'AI-Delivery-Methodology-Checklists.xlsx')
    wb.save(output_file)
    print(f"\n✅ Excel workbook created: {output_file}")
    print(f"   Total sheets: {len(wb.sheetnames)}")

if __name__ == '__main__':
    main()
